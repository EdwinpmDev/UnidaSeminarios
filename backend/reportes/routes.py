from io import BytesIO
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import openpyxl
from flask import Blueprint, current_app, jsonify, make_response, request
from sqlalchemy import extract
from sqlalchemy.orm import selectinload

from extensions import Session
from models import Estudiante, Seminario, UsuarioEvaluador
from utils import aplicar_formato_excel, parsear_jurado

from auth.decorators import admin_requerido, token_requerido

reportes_bp = Blueprint("reportes", __name__)


@reportes_bp.route("/descargar-reporte", methods=["GET"])
@token_requerido
def descargar_reporte():
    session = Session()
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Directorio de alumnos"

        ws.append(['No. Control', 'Nombre de alumno', 'Seminarios activos', 'Seminarios evaluados', 'Proyectos Registrados'])

        estudiantes = session.query(Estudiante).options(
            selectinload(Estudiante.seminarios).selectinload(Seminario.evaluaciones)
        ).all()
        for est in estudiantes:
            activos = 0
            evaluados = 0
            nombres_proyectos = []
            for s in est.seminarios:
                nombres_proyectos.append(s.proyecto)
                evals = s.evaluaciones
                if len(evals) >= 3:
                    evaluados += 1
                else:
                    activos += 1

            proyectos_str = " | ".join(nombres_proyectos) if nombres_proyectos else "Sin proyectos"

            ws.append([est.usuarioAlumno, est.nombre, activos, evaluados, proyectos_str])

        aplicar_formato_excel(ws)

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        respuesta = make_response(out.read())
        respuesta.mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        respuesta.headers["Content-Disposition"] = "attachment;filename=reporte_alumnos.xlsx"
        return respuesta
    finally:
        session.close()


@reportes_bp.route("/descargar-agenda", methods=["GET"])
@token_requerido
def descargar_agenda():
    mes_filtro = request.args.get('mes', 'todos')
    anio_filtro = request.args.get('anio', 'todos')
    programa_filtro = request.args.get('programa', 'todos').strip()
    fase_filtro = request.args.get('fase', 'todos').strip()

    session = Session()
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Agenda filtrada"
        ws.append(['Fecha', 'Hora', 'Lugar', 'Modalidad', 'Estudiante', 'No. Control', 'Tipo de seminario', 'Proyecto'])

        query = session.query(Seminario).join(Estudiante)

        if anio_filtro != 'todos':
            query = query.filter(extract('year', Seminario.fecha) == int(anio_filtro))
        if mes_filtro != 'todos':
            query = query.filter(extract('month', Seminario.fecha) == int(mes_filtro))
        if programa_filtro != 'todos':
            query = query.filter(Estudiante.programa == programa_filtro)
        if fase_filtro != 'todos':
            query = query.filter(Seminario.tipo_seminario == fase_filtro)

        seminarios = query.order_by(Seminario.fecha.asc(), Seminario.hora.asc()).all()

        for s in seminarios:
            if not s.fecha:
                continue

            hora_str = s.hora.strftime('%H:%M') if s.hora else 'Sin hora'
            ws.append([str(s.fecha), hora_str, s.lugar, s.modalidad, s.estudiante.nombre, s.estudiante.usuarioAlumno, s.tipo_seminario, s.proyecto])

        aplicar_formato_excel(ws)

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        respuesta = make_response(out.read())
        respuesta.mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        respuesta.headers["Content-Disposition"] = f"attachment;filename=agenda_{mes_filtro}_{anio_filtro}.xlsx"
        return respuesta
    finally:
        session.close()


@reportes_bp.route("/agenda-paginada", methods=["GET"])
@token_requerido
def agenda_paginada():
    session = Session()
    try:
        mes_filtro = request.args.get('mes', 'todos')
        anio_filtro = request.args.get('anio', 'todos')
        programa_filtro = request.args.get('programa', 'todos').strip()
        fase_filtro = request.args.get('fase', 'todos').strip()
        page = int(request.args.get('page', 1))
        per_page = 15

        query = session.query(Seminario).join(Estudiante)

        if anio_filtro != 'todos':
            query = query.filter(extract('year', Seminario.fecha) == int(anio_filtro))
        if mes_filtro != 'todos':
            query = query.filter(extract('month', Seminario.fecha) == int(mes_filtro))
        if programa_filtro != 'todos':
            query = query.filter(Estudiante.programa == programa_filtro)
        if fase_filtro != 'todos':
            query = query.filter(Seminario.tipo_seminario == fase_filtro)
            
        total_records = query.count()
        has_more = (page * per_page) < total_records

        # Ordenamos los seminarios más recientes primero
        seminarios_bd = query.order_by(Seminario.fecha.asc(), Seminario.hora.asc()).offset((page - 1) * per_page).limit(per_page).all()

        eventos = []
        for sem in seminarios_bd:
            if not sem.fecha:
                continue

            opciones_meses = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
            fecha_str = f"{sem.fecha.day} de {opciones_meses[sem.fecha.month - 1]} de {sem.fecha.year}"

            jurado = parsear_jurado(sem.jurado_texto)

            # Calcular estado de plazo
            estado_plazo = "Activo"
            if sem.fecha and sem.hora:
                fecha_inicio = datetime.combine(sem.fecha, sem.hora).replace(tzinfo=ZoneInfo("America/Mexico_City"))
                fecha_fin = fecha_inicio + timedelta(hours=72)
                ahora = datetime.now(ZoneInfo("America/Mexico_City"))
                
                if ahora < fecha_inicio:
                    estado_plazo = "Pendiente"
                elif ahora > fecha_fin:
                    estado_plazo = "Terminado"

            eventos.append({
                "id_seminario": sem.id,
                "proyecto": sem.proyecto,
                "tipo_seminario": sem.tipo_seminario,
                "lugar": sem.lugar or "No definido",
                "modalidad": sem.modalidad or "Presencial",
                "estado_plazo": estado_plazo,
                "fecha_raw": str(sem.fecha),
                "fecha_bonita": fecha_str,
                "hora": sem.hora.strftime("%H:%M") if sem.hora else "00:00",
                "clave_acceso": sem.clave_acceso,
                "presidente": jurado.get("Presidente", ""),
                "secretario": jurado.get("Secretario", ""),
                "vocal": jurado.get("Vocal", ""),
                "nombre_estudiante": sem.estudiante.nombre,
                "usuarioAlumno": sem.estudiante.usuarioAlumno,
                "programa": sem.estudiante.programa
            })

        return jsonify({
            "success": True,
            "eventos": eventos,
            "has_more": has_more,
            "total": total_records
        }), 200
    except Exception as e:
        current_app.logger.error(f"Error en agenda paginada: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        session.close()


@reportes_bp.route("/descargar-docentes", methods=["GET"])
@admin_requerido
def descargar_docentes():
    session = Session()
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Directorio de docentes"
        ws.append(['ID de base de datos', 'Nombre completo', 'Usuario de acceso'])

        docentes = session.query(UsuarioEvaluador).filter_by(es_admin=False).all()
        for d in docentes:
            ws.append([d.id, d.nombre_completo, d.usuario])

        aplicar_formato_excel(ws)

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        respuesta = make_response(out.read())
        respuesta.mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        respuesta.headers["Content-Disposition"] = "attachment;filename=directorio_docentes.xlsx"
        return respuesta
    finally:
        session.close()
