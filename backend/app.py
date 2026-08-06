import os
import re
import random
import string
import jwt
import io
import logging
import openpyxl
from openpyxl.styles import PatternFill, Font
from io import BytesIO
from logging.handlers import RotatingFileHandler
from functools import wraps
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo

from flask import Flask, request, jsonify, send_from_directory, make_response, redirect
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

from sqlalchemy import create_engine, Column, Integer, String, Float, Text, ForeignKey, DateTime, Date, Time, Boolean, extract
from sqlalchemy.orm import declarative_base, sessionmaker, scoped_session, relationship, selectinload
from sqlalchemy.exc import IntegrityError
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv

# inicializa las variables de entorno
load_dotenv()

# --- CONFIGURACIÓN GENERAL Y SEGURIDAD

DEBUG_MODE = os.getenv("DEBUG", "False").lower() == "true"
JWT_SECRET = os.getenv("JWT_SECRET")

if not JWT_SECRET:
    raise ValueError("CRÍTICO: JWT_SECRET no encontrado en el archivo .env")

FRONTEND_PATH = os.path.join(os.path.dirname(__file__), '../frontend')
app = Flask(__name__, static_folder=None)

# configura el límite de peso por petición para evitar ataques dos (1 mb máximo)
app.config['MAX_CONTENT_LENGTH'] = 1 * 1024 * 1024

# permite peticiones cruzadas para el entorno de desarrollo
CORS(app, resources={
    r"/*": {"origins": ["http://127.0.0.1:5000", "http://localhost:5000", "http://192.168.100.11:5000"]}
})

# previene ataques de fuerza bruta limitando las peticiones por ip
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri="memory://"
)

# crea la carpeta de logs si no existe
os.makedirs('logs', exist_ok=True)

# configura la rotación diaria de bitácoras de auditoría y errores
handler = RotatingFileHandler('logs/unida_auditoria.log', maxBytes=2000000, backupCount=10)
handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
handler.setLevel(logging.INFO if DEBUG_MODE else logging.WARNING)
app.logger.addHandler(handler)
app.logger.setLevel(logging.INFO if DEBUG_MODE else logging.WARNING)

@app.after_request
def aplicar_cabeceras_seguridad(response):
    # inyecta cabeceras http de seguridad en todas las respuestas
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    return response

@app.errorhandler(Exception)
def manejar_error(error):
    if DEBUG_MODE:
        return jsonify({"success": False, "error": str(error)}), 500
    
    # guarda el error en la bitácora sin exponer el código al usuario
    app.logger.error(f"error no controlado detectado: {str(error)}")
    return jsonify({"success": False, "mensaje": "Ocurrió un error interno. El administrador ha sido notificado."}), 500

@app.errorhandler(429)
def demasiados_intentos(error):
    app.logger.warning(f"bloqueo por exceso de peticiones desde ip: {get_remote_address()}")
    return jsonify({"success": False, "mensaje": "Demasiados intentos. Espera un momento y vuelve a intentarlo."}), 429


# --- BASE DE DATOS Y MODELOS ORM

DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise ValueError("CRÍTICO: No se encontró DATABASE_URL en el archivo .env")

engine = create_engine(DATABASE_URL, pool_pre_ping=True, pool_recycle=3600, echo=False)
Base = declarative_base()
Session = scoped_session(sessionmaker(bind=engine))

class UsuarioEvaluador(Base):
    __tablename__ = "usuarios_evaluadores"
    id = Column(Integer, primary_key=True, autoincrement=True)
    usuario = Column(String(50), unique=True, nullable=False, index=True)
    password_hash = Column(String(255), nullable=False)
    nombre_completo = Column(String(100), nullable=False)
    es_admin = Column(Boolean, default=False, nullable=False)

    def verificar_password(self, password):
        return check_password_hash(self.password_hash, password)

class Estudiante(Base):
    __tablename__ = "estudiantes"
    id = Column(Integer, primary_key=True, autoincrement=True)
    usuarioAlumno = Column(String(15), unique=True, nullable=False, index=True)
    password_hash = Column(String(255), nullable=False) 
    nombre = Column(String(100), nullable=False)
    correo = Column(String(150), nullable=False)
    programa = Column(String(50), nullable=False)
    
    seminarios = relationship("Seminario", back_populates="estudiante", cascade="all, delete-orphan")

class Seminario(Base):
    __tablename__ = "seminarios"
    id = Column(Integer, primary_key=True, autoincrement=True)
    estudiante_id = Column(Integer, ForeignKey("estudiantes.id", ondelete="CASCADE"), nullable=False)
    
    clave_acceso = Column(String(20), unique=True, nullable=False)
    clave_presidente = Column(String(20), nullable=True)
    clave_secretario = Column(String(20), nullable=True)
    clave_vocal = Column(String(20), nullable=True)
    
    tipo_seminario = Column(String(50), nullable=False)
    proyecto = Column(Text, nullable=False) 
    estado = Column(String(20), default="Agendado", nullable=False, index=True)
    
    fecha = Column(Date, nullable=True, index=True) 
    hora = Column(Time, nullable=True)
    lugar = Column(String(255), nullable=True)
    modalidad = Column(String(20), nullable=True)
    duracion = Column(String(20), nullable=True)
    jurado_texto = Column(Text, nullable=True)
    observaciones = Column(Text, nullable=True)
    
    estudiante = relationship("Estudiante", back_populates="seminarios")
    evaluaciones = relationship("Evaluacion", back_populates="seminario", cascade="all, delete-orphan")

class Evaluacion(Base):
    __tablename__ = "evaluaciones"
    id = Column(Integer, primary_key=True, autoincrement=True)
    seminario_id = Column(Integer, ForeignKey("seminarios.id", ondelete="CASCADE"), nullable=False)
    evaluador_nombre = Column(String(100), nullable=False)
    evaluador_rol = Column(String(50), nullable=False)
    calificacion_final = Column(Float, nullable=False)
    comentarios = Column(Text, nullable=False)
    fecha_evaluacion = Column(DateTime, default=lambda: datetime.now(ZoneInfo("America/Mexico_City")))
    
    seminario = relationship("Seminario", back_populates="evaluaciones")

Base.metadata.create_all(engine)

# --- UTILIDADES Y MIDDLEWARES

NOMBRE_REGEX = re.compile(r'^[A-Za-zÁÉÍÓÚáéíóúÑñÜü\s]+$')
CORREO_REGEX = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]{2,}$')
NUMERO_CONTROL_REGEX = re.compile(r'^[A-Za-z0-9]{4,}$')
PASSWORD_ESTUDIANTE_REGEX = re.compile(r'^[A-Za-z0-9]{4,}$')

def generar_clave_acceso():
    caracteres = string.ascii_uppercase + string.digits
    return ''.join(random.choice(caracteres) for _ in range(8))

def parsear_jurado(jurado_texto):
    resultado = {"Presidente": "", "Secretario": "", "Vocal": ""}
    if not jurado_texto: return resultado
    for parte in jurado_texto.split('|'):
        if ':' in parte:
            rol, nombre = parte.split(':', 1)
            rol = rol.strip()
            if rol in resultado:
                resultado[rol] = nombre.strip()
    return resultado

def _decodificar_token():
    token = request.cookies.get('unida_token')
    if not token:
        return None, (jsonify({"success": False, "mensaje": "Falta el token de seguridad"}), 401)
    try:
        data = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        return data, None
    except jwt.ExpiredSignatureError:
        return None, (jsonify({"success": False, "mensaje": "Token expirado"}), 401)
    except jwt.InvalidTokenError:
        return None, (jsonify({"success": False, "mensaje": "Token inválido"}), 401)

def token_requerido(f):
    @wraps(f)
    def decorador(*args, **kwargs):
        data, error = _decodificar_token()
        if error: return error
        if 'usuario' not in data:
            return jsonify({"success": False, "mensaje": "Permisos insuficientes."}), 401
        request.usuario_actual = data['usuario']
        return f(*args, **kwargs)
    return decorador

def admin_requerido(f):
    @wraps(f)
    def decorador(*args, **kwargs):
        data, error = _decodificar_token()
        if error: return error
        if 'usuario' not in data or not data.get('is_admin'):
            return jsonify({"success": False, "mensaje": "Se requiere rol de administrador."}), 403
        request.usuario_actual = data['usuario']
        return f(*args, **kwargs)
    return decorador

def csrf_protegido(f):
    # bloquea peticiones que no vengan del mismo dominio para endpoints que modifican datos
    @wraps(f)
    def decorador(*args, **kwargs):
        referer = request.headers.get("Referer")
        if not referer or request.host_url not in referer:
            app.logger.warning(f"intento de csrf bloqueado desde ip: {get_remote_address()}")
            return jsonify({"success": False, "mensaje": "Petición bloqueada por seguridad (CSRF)."}), 403
        return f(*args, **kwargs)
    return decorador

def validar_json(f):
    # fuerza a que todas las peticiones de escritura incluyan el tipo de contenido correcto
    @wraps(f)
    def decorador(*args, **kwargs):
        if not request.is_json:
            return jsonify({"success": False, "mensaje": "La petición debe ser en formato JSON."}), 400
        return f(*args, **kwargs)
    return decorador

def crear_usuario_administrador_inicial():
    # crea al administrador maestro basándose en la configuración de entorno
    session = Session()
    try:
        admin_user = os.getenv("ADMIN_USER")
        admin_pass = os.getenv("ADMIN_PASS")
        if not admin_user or not admin_pass: return

        if not session.query(UsuarioEvaluador).filter_by(usuario=admin_user).first():
            pw_hash = generate_password_hash(admin_pass, method="pbkdf2:sha256", salt_length=16)
            session.add(UsuarioEvaluador(usuario=admin_user, password_hash=pw_hash, nombre_completo="Administrador de seminarios", es_admin=True))
            session.commit()
    except Exception as e:
        session.rollback()
    finally:
        session.close()

def aplicar_formato_excel(ws):
    header_fill = PatternFill(start_color="1B396A", end_color="1B396A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alt_row_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = alt_row_fill

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2


# --- RUTAS DE INTERFAZ HTML

PAGINAS_PROTEGIDAS = {
    'usuario.html': None,
    'index.html': None,
    'portal-alumno.html': None,
    'evaluacion.html': None,
}

@app.route('/')
def index():
    return send_from_directory(FRONTEND_PATH, 'login.html')

@app.route('/<path:path>')
def static_files(path):
    if path in PAGINAS_PROTEGIDAS:
        token = request.cookies.get('unida_token')
        if not token: return redirect('/')
        try:
            data = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
            if PAGINAS_PROTEGIDAS[path] == 'admin' and not data.get('is_admin'):
                return redirect('/')
        except jwt.InvalidTokenError:
            return redirect('/')
    return send_from_directory(FRONTEND_PATH, path)


# --- AUTENTICACIÓN Y SESIONES

@app.route("/verificar-sesion")
def verificar_sesion():
    data, error = _decodificar_token()
    if error: return jsonify({"logueado": False}), 200
    return jsonify({"logueado": True, "usuario": data.get('usuario'), "is_admin": data.get('is_admin', False)})

@app.route("/logout")
def logout():
    respuesta = redirect('/')
    respuesta.set_cookie('unida_token', '', expires=0)
    return respuesta

@app.route("/login", methods=["POST"])
@limiter.limit("5 per minute")
@validar_json
def login():
    session = Session()
    try:
        data = request.get_json()
        usuario, password = data.get("usuario", "").strip(), data.get("password", "").strip()

        if not usuario or len(password) < 8:
            return jsonify({"success": False, "mensaje": "Credenciales inválidas"}), 400

        usuario_db = session.query(UsuarioEvaluador).filter_by(usuario=usuario).first()
        if not usuario_db or not usuario_db.verificar_password(password):
            app.logger.warning(f"intento de login docente fallido para el usuario: {usuario} desde {get_remote_address()}")
            return jsonify({"success": False, "mensaje": "Usuario o contraseña incorrectos"}), 401

        # token de 8 horas
        token_jwt = jwt.encode({"usuario": usuario_db.usuario, "is_admin": usuario_db.es_admin, "exp": datetime.now(timezone.utc) + timedelta(hours=8)}, JWT_SECRET, algorithm="HS256")

        app.logger.info(f"inicio de sesión exitoso: {usuario_db.usuario}")
        respuesta = make_response(jsonify({"success": True, "usuario": usuario_db.nombre_completo, "is_admin": usuario_db.es_admin}))
        respuesta.set_cookie('unida_token', token_jwt, httponly=True, secure=not DEBUG_MODE, samesite='Lax', max_age=28800)
        return respuesta
    finally:
        session.close()

@app.route("/login-evaluador", methods=["POST"])
@limiter.limit("5 per minute")
@validar_json
def login_evaluador():
    session = Session()
    try:
        clave = (request.get_json() or {}).get("seminar_code", "").strip()
        if not clave: return jsonify({"success": False, "mensaje": "Falta clave de acceso"}), 400

        seminario = session.query(Seminario).filter_by(clave_acceso=clave).first()
        if not seminario: 
            app.logger.warning(f"intento de acceso a seminario con clave invalida: {clave} desde {get_remote_address()}")
            return jsonify({"success": False, "mensaje": "Clave incorrecta"}), 401

        token_jwt = jwt.encode({"id_seminario": seminario.id, "rol": "evaluador", "exp": datetime.now(timezone.utc) + timedelta(hours=8)}, JWT_SECRET, algorithm="HS256")
        respuesta = make_response(jsonify({"success": True, "mensaje": "Acceso concedido"}))
        respuesta.set_cookie('unida_token', token_jwt, httponly=True, secure=not DEBUG_MODE, samesite='Lax', max_age=28800)
        return respuesta
    finally:
        session.close()

@app.route("/login-estudiante", methods=["POST"])
@limiter.limit("5 per minute")
@validar_json
def login_estudiante():
    session = Session()
    try:
        data = request.get_json()
        usuarioAlumno, password = data.get("usuario", "").strip(), data.get("password", "").strip()

        estudiante = session.query(Estudiante).filter_by(usuarioAlumno=usuarioAlumno).first()
        if not estudiante or not check_password_hash(estudiante.password_hash, password):
            app.logger.warning(f"intento de login estudiante fallido para control: {usuarioAlumno} desde {get_remote_address()}")
            return jsonify({"success": False, "mensaje": "Número de control o contraseña incorrectos"}), 401

        token_jwt = jwt.encode({"id_estudiante": estudiante.id, "rol": "estudiante", "exp": datetime.now(timezone.utc) + timedelta(hours=8)}, JWT_SECRET, algorithm="HS256")
        respuesta = make_response(jsonify({"success": True, "mensaje": "Login exitoso"}))
        respuesta.set_cookie('unida_token', token_jwt, httponly=True, secure=not DEBUG_MODE, samesite='Lax', max_age=28800)
        return respuesta
    finally:
        session.close()


# --- GESTIÓN CRUD DE ALUMNOS Y SEMINARIOS

@app.route("/registrar-estudiante", methods=["POST"])
@token_requerido
@csrf_protegido
@validar_json
@limiter.limit("10 per minute")
def registrar_estudiante():
    session = Session()
    try:
        data = request.get_json() or {}
        usuarioAlumno, nombre = data.get("usuarioAlumno", "").strip(), data.get("nombre", "").strip()
        password, correo = data.get("password_estudiante", "").strip(), data.get("correo", "").strip()
        presidente = data.get("presidente", "").strip()
        secretario = data.get("secretario", "").strip()
        vocal = data.get("vocal", "").strip()
        lugar = data.get("lugar", "").strip()
        duracion = data.get("duracion", "").strip()

        if not usuarioAlumno or not nombre:
            return jsonify({"success": False, "mensaje": "Faltan datos obligatorios (Control o Nombre)"}), 400
        
        # Verificar si el estudiante ya existe
        estudiante_existente = session.query(Estudiante).filter_by(usuarioAlumno=usuarioAlumno).first()
        
        # Si NO existe, la contraseña es obligatoria
        if not estudiante_existente and not password:
            return jsonify({"success": False, "mensaje": "La contraseña es obligatoria para nuevos alumnos"}), 400
        
        if not NOMBRE_REGEX.match(nombre):
            return jsonify({"success": False, "mensaje": "El nombre solo puede contener letras y espacios"}), 400
        if not NUMERO_CONTROL_REGEX.match(usuarioAlumno):
            return jsonify({"success": False, "mensaje": "El usuario debe tener al menos 8 caracteres/numeros."}), 400
        if password and not PASSWORD_ESTUDIANTE_REGEX.match(password):
            return jsonify({"success": False, "mensaje": "La contraseña debe contener al menos 4 caracteres/numeros"}), 400
        if correo and not CORREO_REGEX.match(correo):
            return jsonify({"success": False, "mensaje": "El correo no es válido (falta un @ o un dominio, ej. .com)"}), 400
        if not presidente or not secretario or not vocal:
            return jsonify({"success": False, "mensaje": "Debes capturar Presidente, Secretario y Vocal del jurado"}), 400
        if not lugar:
            return jsonify({"success": False, "mensaje": "Debes indicar el aula, enlace o modalidad del seminario"}), 400
        if not duracion:
            return jsonify({"success": False, "mensaje": "Debes indicar la duración del seminario"}), 400
        
        try:
            fecha_obj = datetime.strptime(data.get("fecha", ""), "%Y-%m-%d").date()
            hoy_mexico = datetime.now(ZoneInfo("America/Mexico_City")).date()
            if fecha_obj < hoy_mexico:
                return jsonify({"success": False, "mensaje": "No puedes agendar seminarios en el pasado."}), 400
            hora_obj = datetime.strptime(data.get("hora", ""), "%H:%M").time()
        except ValueError:
            return jsonify({"success": False, "mensaje": "Formatos de fecha/hora inválidos."}), 400

        # --- OBTENER O CREAR ESTUDIANTE ---
        estudiante = session.query(Estudiante).filter_by(usuarioAlumno=usuarioAlumno).first()
        if not estudiante:
            # Si no existe, lo creamos (contraseña es obligatoria aqui)
            if not password:
                return jsonify({"success": False, "mensaje": "La contraseña es obligatoria para nuevos alumnos"}), 400
            estudiante = Estudiante(
                usuarioAlumno=usuarioAlumno,
                password_hash=generate_password_hash(password, method="pbkdf2:sha256", salt_length=16),
                nombre=nombre, correo=correo, programa=data.get("programa", "").strip()
            )
            session.add(estudiante)
            session.flush()
        else:
            # Si ya existe, actualizamos
            estudiante.nombre = nombre
            estudiante.correo = correo
            estudiante.programa = data.get("programa", estudiante.programa).strip()
            if password:
                estudiante.password_hash = generate_password_hash(password, method="pbkdf2:sha256", salt_length=16)

        clave_acceso = data.get("clave_acceso", "").strip().upper() or generar_clave_acceso()

        def es_codigo_libre(codigo):
            return not session.query(Seminario).filter(
                (Seminario.clave_acceso == codigo) |
                (Seminario.clave_presidente == codigo) |
                (Seminario.clave_secretario == codigo) |
                (Seminario.clave_vocal == codigo)
            ).first()

        def generar_codigo_unico_global(codigos_usados_en_este_registro):
            for _ in range(10):
                candidato = generar_clave_acceso()
                if candidato not in codigos_usados_en_este_registro and es_codigo_libre(candidato):
                    return candidato
            raise ValueError("No se pudo generar un código único tras varios intentos")

        # Si la clave general ya existe en la base de datos, se reintenta con una nueva
        intentos = 5
        while not es_codigo_libre(clave_acceso) and intentos > 0:
            clave_acceso = generar_clave_acceso()
            intentos -= 1
        if not es_codigo_libre(clave_acceso):
            return jsonify({"success": False, "mensaje": "No se pudo generar una clave de acceso única, intenta de nuevo"}), 500

        codigos_usados = {clave_acceso}
        try:
            clave_presidente = generar_codigo_unico_global(codigos_usados)
            codigos_usados.add(clave_presidente)
            clave_secretario = generar_codigo_unico_global(codigos_usados)
            codigos_usados.add(clave_secretario)
            clave_vocal = generar_codigo_unico_global(codigos_usados)
        except ValueError:
            return jsonify({"success": False, "mensaje": "No se pudieron generar códigos únicos, intenta de nuevo"}), 500
        
        jurado_str = f"Presidente: {data.get('presidente')} | Secretario: {data.get('secretario')} | Vocal: {data.get('vocal')}"
        
        session.add(Seminario(
            estudiante_id=estudiante.id, clave_acceso=clave_acceso,
            clave_presidente=clave_presidente, clave_secretario=clave_secretario, clave_vocal=clave_vocal,
            tipo_seminario=data.get("tipo_seminario", ""), proyecto=data.get("proyecto", "").strip(),
            fecha=fecha_obj, hora=hora_obj, lugar=data.get("lugar", ""), modalidad=data.get("modalidad", ""),
            duracion=data.get("duracion", ""), jurado_texto=jurado_str, observaciones=data.get("observaciones", "")
        ))
        session.commit()
        
        app.logger.info(f"usuario {request.usuario_actual} registro al estudiante {usuarioAlumno}")
        return jsonify({
            "success": True, "mensaje": f"Seminario agendado para {nombre}",
            "clave_acceso": clave_acceso, "clave_presidente": clave_presidente,
            "clave_secretario": clave_secretario, "clave_vocal": clave_vocal
        }), 201
    except IntegrityError:
        session.rollback()
        return jsonify({"success": False, "mensaje": "Número de control duplicado."}), 400
    except Exception as e:
        session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        session.close()

@app.route("/seminario/<int:id_seminario>", methods=["PUT"])
@token_requerido
@csrf_protegido
@validar_json
def editar_seminario(id_seminario):
    session = Session()
    try:
        seminario = session.query(Seminario).filter_by(id=id_seminario).first()
        if not seminario: return jsonify({"success": False, "mensaje": "Seminario no encontrado"}), 404
        
        data = request.get_json() or {}
        
        presidente = data.get("presidente", "").strip()
        secretario = data.get("secretario", "").strip()
        vocal = data.get("vocal", "").strip()
        lugar = data.get("lugar", "").strip()
        duracion = data.get("duracion", "").strip()
        proyecto = data.get("proyecto", "").strip()

        if not presidente or not secretario or not vocal:
            return jsonify({"success": False, "mensaje": "Debes capturar Presidente, Secretario y Vocal del jurado"}), 400
        if not lugar or not duracion or not proyecto:
            return jsonify({"success": False, "mensaje": "Debes indicar el proyecto, lugar y la duración del seminario"}), 400

        try:
            fecha_obj = datetime.strptime(data.get("fecha", ""), "%Y-%m-%d").date()
            hoy_mexico = datetime.now(ZoneInfo("America/Mexico_City")).date()
            if fecha_obj < hoy_mexico:
                return jsonify({"success": False, "mensaje": "La fecha del seminario no puede estar en el pasado"}), 400
            hora_obj = datetime.strptime(data.get("hora", ""), "%H:%M").time()
        except ValueError:
            return jsonify({"success": False, "mensaje": "Fecha u hora inválida"}), 400

        seminario.proyecto = proyecto
        seminario.tipo_seminario = data.get("tipo_seminario")
        seminario.modalidad = data.get("modalidad")
        seminario.lugar = lugar
        seminario.duracion = duracion
        seminario.fecha = fecha_obj
        seminario.hora = hora_obj
        seminario.jurado_texto = f"Presidente: {presidente} | Secretario: {secretario} | Vocal: {vocal}"
        seminario.observaciones = data.get("observaciones", "").strip()

        session.commit()
        app.logger.info(f"usuario {request.usuario_actual} edito el seminario id {id_seminario}")
        return jsonify({"success": True, "mensaje": "Información del seminario actualizada"})
    except Exception as e:
        session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        session.close()

@app.route("/eliminar-estudiante/<int:id_estudiante>", methods=["DELETE"])
@token_requerido
@csrf_protegido
def eliminar_estudiante(id_estudiante):
    session = Session()
    try:
        est = session.query(Estudiante).filter_by(id=id_estudiante).first()
        if not est: return jsonify({"success": False, "mensaje": "No encontrado"}), 404
        
        num_control_respaldo = est.usuarioAlumno
        session.delete(est)
        session.commit()
        
        app.logger.info(f"usuario {request.usuario_actual} elimino al estudiante con control {num_control_respaldo}")
        return jsonify({"success": True, "mensaje": "Registro eliminado del sistema."})
    finally:
        session.close()


# --- OBTENCIÓN DE DATOS (LISTADOS)

@app.route("/estudiantes", methods=["GET"])
@token_requerido
def obtener_estudiantes():
    session = Session()
    try:
        # Obtiene los parametros enviados
        page = int(request.args.get('page', 1))
        per_page = 10
        search = request.args.get('search', '').strip()

        # Prepara la consulta base
        query = session.query(Estudiante).options(
            selectinload(Estudiante.seminarios).selectinload(Seminario.evaluaciones)
        )

        # Si hay texto de búsqueda se aplica el filtro en MySQL
        if search:
            query = query.filter(
                (Estudiante.nombre.ilike(f'%{search}%')) |
                (Estudiante.usuarioAlumno.ilike(f'%{search}%'))
            )

        # Conteo del total sin cargar los datos
        total_records = query.count()
        total_pages = (total_records + per_page - 1) // per_page

        # Extracción de los 10 alumnos que tocan en esta página
        estudiantes = query.order_by(Estudiante.id.desc()).offset((page - 1) * per_page).limit(per_page).all()

        lista = []
        for est in estudiantes:
            sems_list = []
            activos = 0
            for s in est.seminarios:
                evals = s.evaluaciones
                es_evaluado = len(evals) >= 3
                if not es_evaluado:
                    activos += 1
                    promedio_str = f"Pendiente ({len(evals)}/3)"
                else:
                    promedio_str = f"{round(sum(e.calificacion_final for e in evals) / len(evals), 1)} / 100"

                jurado = parsear_jurado(s.jurado_texto)
                
                sems_list.append({
                    "id_seminario": s.id, "proyecto": s.proyecto, "tipo_seminario": s.tipo_seminario,
                    "clave_acceso": s.clave_acceso, "calificacion": promedio_str, "es_evaluado": es_evaluado,
                    "clave_presidente": s.clave_presidente, "presidente": jurado.get("Presidente", ""),
                    "clave_secretario": s.clave_secretario, "secretario": jurado.get("Secretario", ""),
                    "clave_vocal": s.clave_vocal, "vocal": jurado.get("Vocal", ""),
                    "calificacion": promedio_str, "es_evaluado": es_evaluado,
                    "fecha": str(s.fecha) if s.fecha else "", "hora": s.hora.strftime("%H:%M") if s.hora else "",
                    "lugar": s.lugar or "", "modalidad": s.modalidad or ""
                })
            
            sems_list.sort(key=lambda x: x['fecha'], reverse=True)
            lista.append({
                "id_estudiante": est.id, "nombre": est.nombre, "usuarioAlumno": est.usuarioAlumno,
                "correo": est.correo, "programa": est.programa, "seminarios_activos": activos,
                "seminarios": sems_list
            })

        # Retorno de los datos
        return jsonify({
            "success": True,
            "estudiantes": lista,
            "total_pages": total_pages if total_pages > 0 else 1,
            "current_page": page
        }), 200
    except Exception as e:
        session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        session.close()

@app.route("/estudiante/<int:id_estudiante>", methods=["PUT"])
@token_requerido
@csrf_protegido
@validar_json
def editar_solo_estudiante(id_estudiante):
    session = Session()
    try:
        est = session.query(Estudiante).filter_by(id=id_estudiante).first()
        if not est: return jsonify({"success": False, "mensaje": "Estudiante no encontrado"}), 404
        
        data = request.get_json() or {}
        usuarioAlumno = data.get("usuarioAlumno", "").strip()
        nombre = data.get("nombre", "").strip()
        if not usuarioAlumno or not nombre: return jsonify({"success": False, "mensaje": "Control y Nombre obligatorios"}), 400
        
        if usuarioAlumno != est.usuarioAlumno and session.query(Estudiante).filter_by(usuarioAlumno=usuarioAlumno).first():
            return jsonify({"success": False, "mensaje": "Número de control ocupado por otro alumno"}), 400

        if not NUMERO_CONTROL_REGEX.match(usuarioAlumno):
            return jsonify({"success": False, "mensaje": "El número de control debe tener al menos 8 números."}), 400
        if not NOMBRE_REGEX.match(nombre):
            return jsonify({"success": False, "mensaje": "El nombre solo puede contener letras y espacios."}), 400
            
        est.usuarioAlumno = usuarioAlumno
        est.nombre = nombre
        est.correo = data.get("correo", "").strip()
        est.programa = data.get("programa", est.programa).strip()
        pw = data.get("password_estudiante", "").strip()
        if pw: 
            if not PASSWORD_ESTUDIANTE_REGEX.match(pw):
                return jsonify({"success": False, "mensaje": "La contraseña debe contener al menos 6 números y ninguna letra."}), 400
            est.password_hash = generate_password_hash(pw, method="pbkdf2:sha256", salt_length=16)
        
        session.commit()
        return jsonify({"success": True, "mensaje": "Datos del alumno actualizados exitosamente."})
    except Exception as e:
        session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        session.close()

@app.route("/buscar-alumnos-simple", methods=["GET"])
@token_requerido
def buscar_alumnos_simple():
    session = Session()
    try:
        search = request.args.get('search', '').strip()
        page = int(request.args.get('page', 1))
        per_page = 10
        
        query = session.query(Estudiante)
        
        if search:
            query = query.filter(
                (Estudiante.nombre.ilike(f'%{search}%')) | 
                (Estudiante.usuarioAlumno.ilike(f'%{search}%'))
            )
            
        total_records = query.count()
        has_more = (page * per_page) < total_records
            
        estudiantes = query.order_by(Estudiante.id.desc()).offset((page - 1) * per_page).limit(per_page).all()
        
        lista = []
        for est in estudiantes:
            lista.append({
                "id_estudiante": est.id,
                "nombre": est.nombre,
                "usuarioAlumno": est.usuarioAlumno,
                "correo": est.correo,
                "programa": est.programa
            })
            
        return jsonify({
            "success": True, 
            "estudiantes": lista,
            "has_more": has_more,
            "total": total_records
        }), 200
    except Exception as e:
        app.logger.error(f"Error en búsqueda simple: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        session.close()

@app.route("/eliminar-seminario/<int:id_seminario>", methods=["DELETE"])
@token_requerido
@csrf_protegido
def eliminar_seminario(id_seminario):
    session = Session()
    try:
        sem = session.query(Seminario).filter_by(id=id_seminario).first()
        if not sem: return jsonify({"success": False, "mensaje": "No encontrado"}), 404
        session.delete(sem)
        session.commit()
        return jsonify({"success": True, "mensaje": "Seminario eliminado correctamente."})
    finally:
        session.close()

@app.route("/seminario/<int:id_seminario>", methods=["GET"])
@token_requerido
def obtener_seminario_editable(id_seminario):
    session = Session()
    try:
        seminario = session.query(Seminario).filter_by(id=id_seminario).first()
        if not seminario: return jsonify({"success": False, "mensaje": "Seminario no encontrado"}), 404
        estudiante = seminario.estudiante
        jurado = parsear_jurado(seminario.jurado_texto)

        return jsonify({"success": True, "datos": {
            "id_estudiante": estudiante.id, "id_seminario": seminario.id, "usuarioAlumno": estudiante.usuarioAlumno,
            "nombre": estudiante.nombre, "correo": estudiante.correo, "programa": estudiante.programa,
            "proyecto": seminario.proyecto, "tipo_seminario": seminario.tipo_seminario, "modalidad": seminario.modalidad or "",
            "lugar": seminario.lugar or "", "duracion": seminario.duracion or "", "fecha": str(seminario.fecha) if seminario.fecha else "",
            "hora": seminario.hora.strftime("%H:%M") if seminario.hora else "", "presidente": jurado.get("Presidente", ""),
            "secretario": jurado.get("Secretario", ""), "vocal": jurado.get("Vocal", ""), "observaciones": seminario.observaciones or ""
        }}), 200
    finally:
        session.close()

@app.route("/mi-informacion", methods=["GET"])
def mi_informacion():
    token = request.cookies.get('unida_token')
    if not token: return jsonify({"success": False}), 401
    
    session = Session()
    try:
        data = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        if data.get("rol") != "estudiante": return jsonify({"success": False}), 403
        
        estudiante = session.query(Estudiante).filter_by(id=data["id_estudiante"]).first()
        if not estudiante: return jsonify({"success": False}), 404
        
        # Obtenemos TODOS los seminarios del estudiante (del más reciente al más antiguo)
        seminarios_db = session.query(Seminario).filter_by(estudiante_id=estudiante.id).order_by(Seminario.id.desc()).all()
        
        lista_seminarios = []
        for sem in seminarios_db:
            evals = session.query(Evaluacion).filter_by(seminario_id=sem.id).all()
            promedio = round(sum(e.calificacion_final for e in evals) / len(evals), 1) if len(evals) >= 3 else None
            
            jurado_asignado = parsear_jurado(sem.jurado_texto)
            evals_rol = {e.evaluador_rol: e for e in evals}
            estado_jurado = [{"rol": rol, "nombre": jurado_asignado.get(rol) or "Por asignar", "evaluo": rol in evals_rol, "comentarios": evals_rol[rol].comentarios if rol in evals_rol else None} for rol in ["Presidente", "Secretario", "Vocal"]]
            
            lista_seminarios.append({
                "id_seminario": sem.id,
                "proyecto": sem.proyecto,
                "tipo_seminario": sem.tipo_seminario,
                "fecha": str(sem.fecha) if sem.fecha else "Por definir",
                "hora": str(sem.hora) if sem.hora else "Por definir",
                "lugar": sem.lugar or "Por definir",
                "modalidad": sem.modalidad or "Por definir",
                "jurado_texto": sem.jurado_texto or "Por asignar",
                "evaluadores": estado_jurado,
                "promedio": promedio if promedio is not None else f"Pendiente ({len(evals)}/3 evaluaciones)"
            })

        return jsonify({"success": True, "datos": {
            "nombre": estudiante.nombre, 
            "usuarioAlumno": estudiante.usuarioAlumno,
            "correo": estudiante.correo,
            "programa": estudiante.programa,
            "seminarios": lista_seminarios
        }}), 200
    finally:
        session.close()


# --- CÉDULA DE EVALUACIÓN Y RETROALIMENTACIÓN

@app.route("/validar-posicion", methods=["POST"])
@limiter.limit("10 per minute")
@validar_json
def validar_posicion():
    token = request.cookies.get('unida_token')
    if not token: return jsonify({"success": False, "mensaje": "No autorizado"}), 401
    session = Session()
    try:
        token_data = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        if token_data.get("rol") != "evaluador": return jsonify({"success": False, "mensaje": "Acceso denegado"}), 403

        seminario = session.query(Seminario).filter_by(id=token_data["id_seminario"]).first()
        codigo = (request.get_json() or {}).get("codigo_posicion", "").strip().upper()

        mapa_codigos = {"Presidente": seminario.clave_presidente, "Secretario": seminario.clave_secretario, "Vocal": seminario.clave_vocal}
        rol_encontrado = next((r for r, c in mapa_codigos.items() if c and c == codigo), None)

        if not rol_encontrado: return jsonify({"success": False, "mensaje": "Código incorrecto"}), 401
        
        if session.query(Evaluacion).filter_by(seminario_id=seminario.id, evaluador_rol=rol_encontrado).first():
            return jsonify({"success": False, "mensaje": f"El puesto de {rol_encontrado} ya fue evaluado."}), 409
            
        nuevo_token = jwt.encode({"id_seminario": seminario.id, "rol": "evaluador", "rol_evaluador": rol_encontrado, "exp": datetime.now(timezone.utc) + timedelta(hours=8)}, JWT_SECRET, algorithm="HS256")
        respuesta = make_response(jsonify({"success": True, "rol_evaluador": rol_encontrado, "nombre_evaluador": parsear_jurado(seminario.jurado_texto).get(rol_encontrado, "")}))
        respuesta.set_cookie('unida_token', nuevo_token, httponly=True, secure=not DEBUG_MODE, samesite='Lax', max_age=28800)
        return respuesta
    finally:
        session.close()

@app.route("/guardar-evaluacion", methods=["POST"])
@limiter.limit("10 per minute")
@csrf_protegido
@validar_json
def guardar_evaluacion():
    token = request.cookies.get('unida_token')
    if not token: return jsonify({"success": False, "mensaje": "No autorizado"}), 401
    session = Session()
    try:
        token_data = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        id_seminario, evaluador_rol = token_data.get("id_seminario"), token_data.get("rol_evaluador")
        
        if not evaluador_rol:
            return jsonify({"success": False, "mensaje": "Debes validar tu posición primero."}), 403

        if session.query(Evaluacion).filter_by(seminario_id=id_seminario, evaluador_rol=evaluador_rol).first():
            return jsonify({"success": False, "mensaje": "No se puede evaluar dos veces."}), 409
        
        data = request.get_json() or {}
        respuestas = []
        for i in range(1, 13):
            valor = data.get(f"P{i}")
            if valor is None:
                return jsonify({"success": False, "mensaje": f"Falta responder la pregunta P{i}"}), 400
            respuestas.append(float(valor))

        calif_final = round(((sum(respuestas[0:8]) + (sum(respuestas[8:12]) * 2.0)) / 120.0) * 100, 1)

        session.add(Evaluacion(
            seminario_id=id_seminario, evaluador_nombre=data.get("evaluador_nombre"),
            evaluador_rol=evaluador_rol, calificacion_final=calif_final, comentarios=data.get("comentarios", "").strip()
        ))
        session.commit()

        app.logger.info(f"evaluacion registrada para seminario {id_seminario} por el rol de {evaluador_rol}")
        respuesta = make_response(jsonify({"success": True, "calificacion": calif_final}))
        respuesta.set_cookie('unida_token', '', expires=0) 
        return respuesta, 201
    finally:
        session.close()

@app.route("/datos-evaluacion", methods=["GET"])
def datos_evaluacion():
    token = request.cookies.get('unida_token')
    if not token: return jsonify({"success": False, "mensaje": "No autorizado"}), 401
    session = Session()
    try:
        data = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        seminario = session.query(Seminario).filter_by(id=data["id_seminario"]).first()
        roles_ya_evaluados = [e.evaluador_rol for e in session.query(Evaluacion).filter_by(seminario_id=seminario.id).all()]
        
        rol_actual = data.get("rol_evaluador")
        return jsonify({"success": True, "datos": {
            "id_seminario": seminario.id, "nombre_estudiante": seminario.estudiante.nombre,
            "proyecto": seminario.proyecto, "programa": seminario.estudiante.programa,
            "tipo_seminario": seminario.tipo_seminario, "roles_evaluados": roles_ya_evaluados,
            "rol_evaluador": rol_actual, "nombre_evaluador": parsear_jurado(seminario.jurado_texto).get(rol_actual, "") if rol_actual else ""
        }}), 200
    finally:
        session.close()

@app.route("/retroalimentacion/<int:id_seminario>", methods=["GET"])
@token_requerido
def obtener_retroalimentacion(id_seminario):
    session = Session()
    try:
        seminario = session.query(Seminario).filter_by(id=id_seminario).first()
        if not seminario: return jsonify({"success": False}), 404

        evaluaciones = session.query(Evaluacion).filter_by(seminario_id=id_seminario).all()
        roles_eval = {e.evaluador_rol for e in evaluaciones}
        jurado_asignado = parsear_jurado(seminario.jurado_texto)

        return jsonify({
            "success": True, "estudiante": seminario.estudiante.nombre,
            "proyecto": seminario.proyecto, "clave_acceso": seminario.clave_acceso,
            "observaciones": seminario.observaciones,
            "codigos_posicion": {
                rol: {"codigo": getattr(seminario, f"clave_{rol.lower()}") if rol not in roles_eval else None, "nombre": jurado_asignado.get(rol, "")}
                for rol in ["Presidente", "Secretario", "Vocal"]
            },
            "evaluaciones": [{"rol": e.evaluador_rol, "nombre": e.evaluador_nombre, "calificacion": e.calificacion_final, "comentarios": e.comentarios} for e in evaluaciones]
        }), 200
    finally:
        session.close()

# --- MÓDULO DE GESTIÓN DE DOCENTES

@app.route("/docentes", methods=["GET"])
@admin_requerido
def obtener_docentes():
    session = Session()
    try:
        page = int(request.args.get('page', 1))
        per_page = 10
        search = request.args.get('search', '').strip()

        query = session.query(UsuarioEvaluador).filter_by(es_admin=False)

        if search:
            query = query.filter(
                (UsuarioEvaluador.nombre_completo.ilike(f'%{search}%')) |
                (UsuarioEvaluador.usuario.ilike(f'%{search}%'))
            )

        total_records = query.count()
        total_pages = (total_records + per_page - 1) // per_page

        docentes = query.order_by(UsuarioEvaluador.id.desc()).offset((page - 1) * per_page).limit(per_page).all()

        lista = [{"id": d.id, "usuario": d.usuario, "nombre_completo": d.nombre_completo} for d in docentes]

        return jsonify({
            "success": True,
            "docentes": lista,
            "total_pages": total_pages if total_pages > 0 else 1,
            "current_page": page
        }), 200
    except Exception as e:
        session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        session.close()

@app.route("/registrar-docente", methods=["POST"])
@admin_requerido
@csrf_protegido
@validar_json
def registrar_docente():
    session = Session()
    try:
        data = request.get_json()
        usuario, password, nombre = data.get("usuario", "").strip(), data.get("password", "").strip(), data.get("nombre_completo", "").strip()

        if not usuario or not nombre:
            return jsonify({"success": False, "mensaje": "Nombre y usuario son obligatorios"}), 400
        if len(password) < 8: return jsonify({"success": False, "mensaje": "Contraseña muy corta"}), 400
        if session.query(UsuarioEvaluador).filter_by(usuario=usuario).first():
            return jsonify({"success": False, "mensaje": "Usuario ocupado"}), 400

        session.add(UsuarioEvaluador(usuario=usuario, password_hash=generate_password_hash(password, method="pbkdf2:sha256", salt_length=16), nombre_completo=nombre))
        session.commit()
        
        app.logger.info(f"usuario {request.usuario_actual} creo al docente {usuario}")
        return jsonify({"success": True, "mensaje": f"Docente registrado."}), 201
    finally:
        session.close()

@app.route("/editar-docente/<int:id_docente>", methods=["PUT"])
@admin_requerido
@csrf_protegido
@validar_json
def editar_docente(id_docente):
    session = Session()
    try:
        docente = session.query(UsuarioEvaluador).filter_by(id=id_docente).first()
        if not docente:
            return jsonify({"success": False, "mensaje": "Docente no encontrado"}), 404
        if docente.es_admin: return jsonify({"success": False, "mensaje": "No puedes editar a un administrador"}), 403

        data = request.get_json() or {}
        usuario_nuevo = data.get("usuario", "").strip()
        nombre_nuevo = data.get("nombre_completo", "").strip()

        if not usuario_nuevo or not nombre_nuevo:
            return jsonify({"success": False, "mensaje": "Nombre y usuario son obligatorios"}), 400
        
        if usuario_nuevo != docente.usuario and session.query(UsuarioEvaluador).filter_by(usuario=usuario_nuevo).first():
            return jsonify({"success": False, "mensaje": "Usuario ocupado"}), 400

        docente.nombre_completo, docente.usuario = data.get("nombre_completo", "").strip(), usuario_nuevo
        if data.get("password"): docente.password_hash = generate_password_hash(data.get("password").strip(), method="pbkdf2:sha256", salt_length=16)

        session.commit()
        return jsonify({"success": True, "mensaje": "Docente actualizado."})
    finally:
        session.close()

@app.route("/eliminar-docente/<int:id_docente>", methods=["DELETE"])
@admin_requerido
@csrf_protegido
def eliminar_docente(id_docente):
    session = Session()
    try:
        docente = session.query(UsuarioEvaluador).filter_by(id=id_docente).first()
        if not docente:
            return jsonify({"success": False, "mensaje": "Docente no encontrado"}), 404
        if docente.es_admin: return jsonify({"success": False, "mensaje": "No se puede eliminar a un administrador"}), 403
        
        usuario_respaldo = docente.usuario
        session.delete(docente)
        session.commit()
        
        app.logger.info(f"usuario {request.usuario_actual} elimino al docente {usuario_respaldo}")
        return jsonify({"success": True, "mensaje": "Docente eliminado."})
    finally:
        session.close()

# --- MÓDULO DE EXPORTACIÓN A EXCEL

@app.route("/descargar-reporte", methods=["GET"])
@token_requerido
def descargar_reporte():
    session = Session()
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Directorio de alumnos"
        
        ws.append(['No. Control', 'Nombre de alumno', 'Seminarios activos', 'Seminarios evaluados', 'Proyectos Registrados'])
        
        estudiantes = session.query(Estudiante).options(
            selectinload(Estudiante.seminarios).selectinload(Seminario.evaluaciones)
        ).all()
        for est in estudiantes:
            activos = 0
            evaluados = 0
            nombres_proyectos = []
            for s in est.seminarios:
                nombres_proyectos.append(s.proyecto)
                evals = s.evaluaciones
                if len(evals) >= 3:
                    evaluados += 1
                else:
                    activos += 1
            
            proyectos_str = " | ".join(nombres_proyectos) if nombres_proyectos else "Sin proyectos"
            
            ws.append([est.usuarioAlumno, est.nombre, activos, evaluados, proyectos_str])
            
        aplicar_formato_excel(ws)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        respuesta = make_response(out.read())
        respuesta.mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        respuesta.headers["Content-Disposition"] = "attachment;filename=reporte_alumnos.xlsx"
        return respuesta
    finally:
        session.close()

@app.route("/descargar-agenda", methods=["GET"])
@token_requerido
def descargar_agenda():
    mes_filtro = request.args.get('mes', 'todos')
    anio_filtro = request.args.get('anio', 'todos')
    
    session = Session()
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Agenda filtrada"
        ws.append(['Fecha', 'Hora', 'Lugar', 'Modalidad', 'Estudiante', 'No. Control', 'Tipo de seminario', 'Proyecto'])

        seminarios = session.query(Seminario).order_by(Seminario.fecha.asc(), Seminario.hora.asc()).all()
        
        for s in seminarios:
            if not s.fecha: continue
            
            s_mes = f"{s.fecha.month:02d}"
            s_anio = str(s.fecha.year)
            
            if mes_filtro != 'todos' and s_mes != mes_filtro: continue
            if anio_filtro != 'todos' and s_anio != anio_filtro: continue
            
            hora_str = s.hora.strftime('%H:%M') if s.hora else 'Sin hora'
            ws.append([str(s.fecha), hora_str, s.lugar, s.modalidad, s.estudiante.nombre, s.estudiante.usuarioAlumno, s.tipo_seminario, s.proyecto])
            
        aplicar_formato_excel(ws)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        respuesta = make_response(out.read())
        respuesta.mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        respuesta.headers["Content-Disposition"] = f"attachment;filename=agenda_{mes_filtro}_{anio_filtro}.xlsx"
        return respuesta
    finally:
        session.close()

@app.route("/agenda-paginada", methods=["GET"])
@token_requerido
def agenda_paginada():
    session = Session()
    try:
        mes_filtro = request.args.get('mes', 'todos')
        anio_filtro = request.args.get('anio', 'todos')
        page = int(request.args.get('page', 1))
        per_page = 15

        query = session.query(Seminario).join(Estudiante)
        
        if anio_filtro != 'todos':
            query = query.filter(extract('year', Seminario.fecha) == int(anio_filtro))
        if mes_filtro != 'todos':
            query = query.filter(extract('month', Seminario.fecha) == int(mes_filtro))
            
        total_records = query.count()
        has_more = (page * per_page) < total_records
        
        # Ordenamos los seminarios más recientes primero
        seminarios_bd = query.order_by(Seminario.fecha.asc(), Seminario.hora.asc()).offset((page - 1) * per_page).limit(per_page).all()
        
        eventos = []
        for sem in seminarios_bd:
            if not sem.fecha: continue
            
            opciones_meses = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
            fecha_str = f"{sem.fecha.day} de {opciones_meses[sem.fecha.month - 1]} de {sem.fecha.year}"
            
            eventos.append({
                "id_seminario": sem.id,
                "proyecto": sem.proyecto,
                "tipo_seminario": sem.tipo_seminario,
                "lugar": sem.lugar or "No definido",
                "modalidad": sem.modalidad or "Presencial",
                "fecha_raw": str(sem.fecha),
                "fecha_bonita": fecha_str,
                "hora": sem.hora.strftime("%H:%M") if sem.hora else "00:00",
                "clave_acceso": sem.clave_acceso,
                "presidente": sem.clave_presidente,
                "secretario": sem.clave_secretario,
                "vocal": sem.clave_vocal,
                "nombre_estudiante": sem.estudiante.nombre,
                "usuarioAlumno": sem.estudiante.usuarioAlumno,
                "programa": sem.estudiante.programa
            })
            
        return jsonify({
            "success": True, 
            "eventos": eventos,
            "has_more": has_more,
            "total": total_records
        }), 200
    except Exception as e:
        app.logger.error(f"Error en agenda paginada: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        session.close()

@app.route("/descargar-docentes", methods=["GET"])
@admin_requerido
def descargar_docentes():
    session = Session()
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Directorio de docentes"
        ws.append(['ID de base de datos', 'Nombre completo', 'Usuario de acceso'])
        
        docentes = session.query(UsuarioEvaluador).filter_by(es_admin=False).all()
        for d in docentes:
            ws.append([d.id, d.nombre_completo, d.usuario])
            
        aplicar_formato_excel(ws)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        respuesta = make_response(out.read())
        respuesta.mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        respuesta.headers["Content-Disposition"] = "attachment;filename=directorio_docentes.xlsx"
        return respuesta
    finally:
        session.close()

if __name__ == "__main__":
    crear_usuario_administrador_inicial()
    # Para producción, usar Gunicorn en lugar de app.run()
    # Pero si usas app.run(), quita el debug
    app.run(host="0.0.0.0", port=5000, debug=False)