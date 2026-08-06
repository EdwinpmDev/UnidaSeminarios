import random
import re
import string

from openpyxl.styles import Font, PatternFill

# --- EXPRESIONES REGULARES DE VALIDACIÓN ---
NOMBRE_REGEX = re.compile(r'^[A-Za-zÁÉÍÓÚáéíóúÑñÜü\s]+$')
CORREO_REGEX = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]{2,}$')
NUMERO_CONTROL_REGEX = re.compile(r'^[A-Za-z0-9]{4,}$')
PASSWORD_ESTUDIANTE_REGEX = re.compile(r'^[A-Za-z0-9]{4,}$')


def generar_clave_acceso():
    caracteres = string.ascii_uppercase + string.digits
    return ''.join(random.choice(caracteres) for _ in range(8))


def parsear_jurado(jurado_texto):
    resultado = {"Presidente": "", "Secretario": "", "Vocal": ""}
    if not jurado_texto:
        return resultado
    for parte in jurado_texto.split('|'):
        if ':' in parte:
            rol, nombre = parte.split(':', 1)
            rol = rol.strip()
            if rol in resultado:
                resultado[rol] = nombre.strip()
    return resultado


def aplicar_formato_excel(ws):
    header_fill = PatternFill(start_color="1B396A", end_color="1B396A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alt_row_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = alt_row_fill

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
